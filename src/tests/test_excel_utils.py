"""Unit tests for the create_excel_table function."""
import unittest
from openpyxl import Workbook
from src.main.excel_utils import create_excel_table


class TestCreateExcelTable(unittest.TestCase):
    """Test cases for create_excel_table function."""

    def setUp(self):
        """Set up test fixtures."""
        self.workbook = Workbook()

    def test_create_excel_table_happy_path(self):
        """Test create_excel_table with explicit sheet name."""

        # Create a new worksheet named 'TestSheet'
        self.workbook.create_sheet('TestSheet')

        # Fetch the worksheet
        worksheet = self.workbook['TestSheet']

        # Call the function with real objects
        create_excel_table(self.workbook, 'TestSheet')

        # Fetch the table from the worksheet
        # Accessing _tables is necessary for this test to verify table creation.
        # pylint: disable=protected-access
        table = worksheet._tables.get("Table1", None)

        self.assertIsNotNone(table)
        self.assertEqual(worksheet.title, "TestSheet")
        self.assertEqual(table.displayName, "Table1")

    def test_create_excel_table_default_sheet(self):
        """Test create_excel_table with default sheet name."""
        # Remove the default sheet and create a new worksheet named 'Sheet1'
        self.workbook.remove(self.workbook.active)
        self.workbook.create_sheet('Sheet1')

        # Call the function with real objects
        create_excel_table(self.workbook)

        # Fetch the worksheet
        worksheet = self.workbook['Sheet1']

        # Fetch the table from the worksheet
        # Accessing _tables is necessary for this test to verify table creation.
        # pylint: disable=protected-access
        table = worksheet._tables.get("Table1", None)

        self.assertIsNotNone(table)
        self.assertEqual(worksheet.title, "Sheet1")
        self.assertEqual(table.displayName, "Table1")


if __name__ == '__main__':
    unittest.main()
